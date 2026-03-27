"""
사학진흥재단 법인교비 결산 xlsx → data/건축비.json 변환 스크립트

건축비 = 건물매입비[1263] + 건설가계정[1270] (천원 단위 → ×1000 → 원)

사용법:
    # 단일 파일
    python convert_건축비.py "공시자료/건축비test/법인일반회계 및 교비회계 결산(2024회계연도).xlsx"

    # 폴더 일괄 처리
    python convert_건축비.py "공시자료/건축비test/"

지원 파일 형식:
    - {연도}회계연도 사립대학 법인교비 결산.xlsx       (2015~2020)
    - 법인일반 및 교비회계(통합) 결산({연도}회계연도).xlsx  (2021~2022)
    - 법인일반회계 및 교비회계 결산({연도}회계연도).xlsx    (2023~)

출력 JSON 구조:
    [{"기준연도": 2024, "공시연도": 2025, "기준대학명": "가야대학교", "건축비_원": 532176257}, ...]

    기준연도 = 파일명의 회계연도
    공시연도 = 기준연도 + 1
"""

import sys
import json
import re
import os
from pathlib import Path

# Windows 콘솔 UTF-8 출력
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

try:
    import openpyxl
except ImportError:
    print("openpyxl이 필요합니다: pip install openpyxl")
    sys.exit(1)

BASE_DIR = Path(__file__).parent
OUTPUT_PATH = BASE_DIR / 'data' / '건축비.json'
기준대학_PATH = BASE_DIR / 'data' / '기준대학.json'


# ── 기준대학 매핑 로드 ───────────────────────────────────────
def load_base_univ_map() -> dict[str, str]:
    """기준대학.json → {대학명: 기준대학명} 역방향 맵"""
    with open(기준대학_PATH, encoding='utf-8') as f:
        entries = json.load(f)
    return {e['대학명']: e['기준대학명'] for e in entries if e.get('대학명')}


# ── 파일명에서 회계연도 추출 ─────────────────────────────────
def extract_year(filename: str) -> int | None:
    """파일명에서 4자리 연도 추출.

    예)
        2024회계연도 사립대학 법인교비 결산.xlsx → 2024
        법인일반회계 및 교비회계 결산(2024회계연도).xlsx → 2024
    """
    m = re.search(r'(\d{4})회계연도', filename)
    return int(m.group(1)) if m else None


# ── xlsx 파싱 ────────────────────────────────────────────────
def parse_xlsx(xlsx_path: Path, base_univ_map: dict[str, str]) -> tuple[list[dict], list[str]]:
    """xlsx 1개 파일 → (레코드 목록, 미매핑 대학명 목록)"""
    기준연도 = extract_year(xlsx_path.name)
    if 기준연도 is None:
        print(f"  [SKIP] 파일명에서 연도를 찾을 수 없음: {xlsx_path.name}")
        return [], []
    공시연도 = 기준연도 + 1

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # 시트 선택: '자금' 포함 시트 우선
    target_sheet = next((s for s in wb.sheetnames if '자금' in s), wb.sheetnames[0])
    ws = wb[target_sheet]

    # 헤더행 탐색: '학교명' 셀이 있는 행 (최대 10행 탐색)
    header_row_idx = None
    headers: list = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if any(v and '학교명' in str(v) for v in row):
            header_row_idx = r_idx
            headers = list(row)
            break

    if header_row_idx is None:
        print(f"  [SKIP] 헤더행(학교명 포함)을 찾을 수 없음: {xlsx_path.name}")
        wb.close()
        return [], []

    # 컬럼 인덱스 탐색
    col_학교명 = next((i for i, h in enumerate(headers) if h and '학교명' in str(h)), None)
    col_회계   = next((i for i, h in enumerate(headers) if h and str(h).strip() == '회계'), None)
    col_건물   = next((i for i, h in enumerate(headers) if h and '[1263]' in str(h)), None)
    col_건설   = next((i for i, h in enumerate(headers) if h and '[1270]' in str(h)), None)

    if col_학교명 is None or col_회계 is None:
        print(f"  [SKIP] 필수 컬럼(학교명/회계) 미발견: {xlsx_path.name}")
        wb.close()
        return [], []

    missing_cols = []
    if col_건물 is None:
        missing_cols.append('건물매입비[1263]')
    if col_건설 is None:
        missing_cols.append('건설가계정[1270]')
    if missing_cols:
        print(f"  [경고] 컬럼 미발견 → 0으로 처리: {', '.join(missing_cols)} ({xlsx_path.name})")

    # 데이터 읽기
    records: list[dict] = []
    unmatched: list[str] = []

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        # 교비 행만
        if not row[col_회계] or str(row[col_회계]).strip() != '교비':
            continue

        학교명_raw = str(row[col_학교명] or '').strip()
        if not 학교명_raw:
            continue

        # 기준대학명 정규화: 직접 조회 → (구.XXX) 등 괄호 제거 후 재시도
        기준대학명 = base_univ_map.get(학교명_raw)
        if 기준대학명 is None:
            stripped = re.sub(r'\s*\(.*?\)\s*$', '', 학교명_raw).strip()
            기준대학명 = base_univ_map.get(stripped)
        if 기준대학명 is None:
            if 학교명_raw not in unmatched:
                unmatched.append(학교명_raw)

        건물 = float(row[col_건물] or 0) if col_건물 is not None else 0.0
        건설 = float(row[col_건설] or 0) if col_건설 is not None else 0.0
        건축비_원 = round((건물 + 건설) * 1000)  # 천원 → 원

        records.append({
            '기준연도': 기준연도,
            '공시연도': 공시연도,
            '기준대학명': 기준대학명,   # 매핑 실패 시 null — 누락하지 않고 포함
            '원본학교명': 학교명_raw,
            '건축비_원': 건축비_원,
        })

    wb.close()
    return records, unmatched


# ── JSON upsert ──────────────────────────────────────────────
def upsert_json(new_records: list[dict]) -> None:
    """data/건축비.json에 기준연도 기준 upsert (같은 연도 재실행 시 덮어쓰기)"""
    if not new_records:
        return

    # 기존 데이터 로드
    existing: list[dict] = []
    if OUTPUT_PATH.exists():
        with open(OUTPUT_PATH, encoding='utf-8') as f:
            existing = json.load(f)

    # 새 데이터의 (기준연도, 원본학교명) 키 집합 — 기준대학명이 null일 수 있어 원본으로 식별
    new_keys = {(r['기준연도'], r['원본학교명']) for r in new_records}

    # 기존 데이터에서 같은 키 제거 후 병합
    merged = [r for r in existing if (r['기준연도'], r.get('원본학교명', r['기준대학명'])) not in new_keys]
    merged.extend(new_records)

    # 정렬: 기준연도 오름차순, 원본학교명 오름차순 (기준대학명이 null일 수 있음)
    merged.sort(key=lambda r: (r['기준연도'], r['원본학교명']))

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)


# ── 진입점 ───────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("사용법: python convert_건축비.py <xlsx파일 또는 폴더>")
        sys.exit(1)

    target = Path(sys.argv[1])
    if not target.exists():
        print(f"경로를 찾을 수 없습니다: {target}")
        sys.exit(1)

    # 처리할 파일 목록
    if target.is_dir():
        files = sorted(target.glob('*.xlsx'))
    else:
        files = [target]

    if not files:
        print("xlsx 파일이 없습니다.")
        sys.exit(0)

    # 기준대학 맵 로드
    if not 기준대학_PATH.exists():
        print(f"기준대학.json이 없습니다: {기준대학_PATH}")
        sys.exit(1)
    base_univ_map = load_base_univ_map()
    print(f"기준대학 매핑 {len(base_univ_map)}개 로드 완료\n")

    all_records: list[dict] = []
    all_unmatched: list[str] = []

    for fp in files:
        연도 = extract_year(fp.name)
        연도_str = f"{연도}년" if 연도 else "연도불명"
        print(f"[{연도_str}] {fp.name}")
        records, unmatched = parse_xlsx(fp, base_univ_map)
        print(f"  → 교비 매칭 {len(records)}개 / 미매핑 {len(unmatched)}개")
        all_records.extend(records)
        all_unmatched.extend(u for u in unmatched if u not in all_unmatched)

    if not all_records:
        print("\n저장할 데이터가 없습니다.")
        return

    upsert_json(all_records)
    print(f"\n[완료] data/건축비.json 저장 완료 (총 {len(all_records)}개 레코드 추가/갱신)")

    if all_unmatched:
        print(f"\n[경고] 기준대학.json 미매핑 대학명 ({len(all_unmatched)}개):")
        for name in sorted(all_unmatched):
            print(f"   - {name}")
        print("   → admin.html 기준대학 매핑 탭에서 별칭 추가 후 재실행하세요.")


if __name__ == '__main__':
    main()
