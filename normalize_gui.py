"""
대학알리미 공시 파일 정제 + JSON 누적 GUI

흐름:
  1. 입력 폴더에서 xlsx 파일 정제 (병합해제, 헤더 flatten)
  2. 필드명 변경 감지 → 확인 팝업 (기존 필드에 매핑 or 새 필드로 추가)
  3. CSV / Excel 저장 (선택, 검증용)
  4. 항목별 JSON에 연도 데이터 누적 (분석툴용)

파일 구조:
  normalize_gui.py      ← 이 파일
  field_mapping.json    ← 필드 매핑 (자동 생성/관리)
  data/
    전임교원_1인당_학생수_및_확보율.json   ← 항목별 누적 JSON
    재적학생수.json
    ...

사용법:
  pip install pandas openpyxl
  python normalize_gui.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os, re, json
import openpyxl
import pandas as pd
from pathlib import Path
from copy import deepcopy

# ── 대학알리미 xlsx 파일 내부 XML 버그 대응 ──────────────
# (1) 'cumstomWidth'(오타) 키워드 무시
_orig_col_init = openpyxl.worksheet.dimensions.ColumnDimension.__init__
def _patched_col_init(self, worksheet, *args, **kwargs):
    kwargs.pop('cumstomWidth', None)
    _orig_col_init(self, worksheet, *args, **kwargs)
openpyxl.worksheet.dimensions.ColumnDimension.__init__ = _patched_col_init

# (2) 존재하지 않는 drawing XML 참조 무시
#     xlsx 내부 ZIP에 drawing1.xml이 없는데 참조만 있는 경우 KeyError 발생
from openpyxl.reader import drawings as _drawings_mod
_orig_find_images = _drawings_mod.find_images
def _patched_find_images(archive, path):
    try:
        return _orig_find_images(archive, path)
    except KeyError:
        return [], []   # drawing 파일 누락 → 빈 결과로 계속
_drawings_mod.find_images = _patched_find_images
# excel reader도 같은 함수를 직접 import해서 쓰므로 거기도 교체
from openpyxl.reader import excel as _excel_mod
_excel_mod.find_images = _patched_find_images
# ─────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────
# 설정
# ─────────────────────────────────────────────────────
YEAR_PATTERN     = re.compile(r"^(19|20)\d{2}(\s*년\s*(상|하)반기)?$")
YEAR_IN_FILENAME = re.compile(r"^\d{4}년[_\s]*")   # 파일명 앞 연도 제거용
SCRIPT_DIR       = Path(__file__).parent
MAPPING_FILE     = SCRIPT_DIR / "field_mapping.json"
JSON_DIR         = SCRIPT_DIR / "data"


# ─────────────────────────────────────────────────────
# 정제 로직
# ─────────────────────────────────────────────────────

def clean_str(val) -> str:
    if val is None:
        return ""
    return str(val).replace("\n", " ").strip()


def pick_sheet(wb):
    """시트 선택: Sheet1 → raw 포함 → empty 포함 → 첫 번째"""
    names = wb.sheetnames
    for name in names:
        if name.lower() == "sheet1":
            return wb[name], name
    for name in names:
        if "raw" in name.lower():
            return wb[name], name
    for name in names:
        if "empty" in name.lower():
            return wb[name], name
    return wb[names[0]], names[0]


def detect_data_start(ws) -> int:
    """병합된 상태에서도 A열 연도값 첫 등장 행 반환"""
    merge_map = {}
    for mr in ws.merged_cells.ranges:
        val = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merge_map[(r, c)] = val

    for row_idx in range(1, ws.max_row + 1):
        raw = merge_map.get((row_idx, 1), ws.cell(row_idx, 1).value)
        if YEAR_PATTERN.match(clean_str(raw)):
            return row_idx
    raise ValueError("A열에서 연도값을 찾을 수 없습니다.")


def unmerge_and_fill(ws):
    for mr in list(ws.merged_cells.ranges):
        top_left = ws.cell(mr.min_row, mr.min_col).value
        ws.unmerge_cells(str(mr))
        for row in ws.iter_rows(min_row=mr.min_row, max_row=mr.max_row,
                                 min_col=mr.min_col, max_col=mr.max_col):
            for cell in row:
                cell.value = top_left


def flatten_headers(ws, header_rows: list) -> list:
    flat = []
    for col in range(1, ws.max_column + 1):
        parts = []
        for row in header_rows:
            v = clean_str(ws.cell(row, col).value)
            if v and v not in parts:
                parts.append(v)
        flat.append("_".join(parts) if parts else f"col{col}")
    return flat


def extract_df(filepath: str) -> tuple:
    """xlsx → (df, sheet_name, header_rows, data_start) 반환"""
    wb   = openpyxl.load_workbook(filepath)
    ws, sheet_name = pick_sheet(wb)

    data_start = detect_data_start(ws)   # 병합 해제 전에 먼저
    unmerge_and_fill(ws)

    header_rows = []
    for r in range(1, data_start):
        vals        = [clean_str(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        unique_vals = set(v for v in vals if v)
        if len(unique_vals) >= 2:
            header_rows.append(r)

    if not header_rows:
        raise ValueError("헤더 행을 찾을 수 없습니다.")

    headers = flatten_headers(ws, header_rows)

    data = [list(row) for row in ws.iter_rows(min_row=data_start, values_only=True)
            if any(v is not None for v in row)]

    df = pd.DataFrame(data, columns=headers)
    return df, sheet_name, header_rows, data_start


HALF_YEAR_PATTERN = re.compile(r"^((19|20)\d{2})\s*년\s*(상|하)반기")

def normalize_half_year(df: "pd.DataFrame") -> "pd.DataFrame":
    """기준연도 컬럼에 '20XX 년 상반기 / 하반기' 형식이 있으면:
    - 하반기 행 제거 (연중 중간 집계라 불완전)
    - 상반기 행의 기준연도를 4자리 정수로 정규화
    해당 형식이 없으면 df를 그대로 반환.
    """
    if df.empty:
        return df
    year_col = df.columns[0]

    def _match(val):
        return HALF_YEAR_PATTERN.match(str(val).strip())

    if not df[year_col].apply(_match).any():
        return df   # 이 항목에는 반기 표기 없음

    # 하반기 제거
    mask_ha = df[year_col].apply(
        lambda v: bool(m := HALF_YEAR_PATTERN.match(str(v).strip())) and m.group(3) == '하'
    )
    df = df[~mask_ha].copy()

    # 상반기 → 4자리 정수
    def _norm(val):
        m = HALF_YEAR_PATTERN.match(str(val).strip())
        return int(m.group(1)) if m else val

    df[year_col] = df[year_col].apply(_norm)
    return df


def item_key_from_filename(filename: str) -> str:
    """파일명에서 연도 제거 → 항목 키
    예) '2024년__대학_6-나-(1)_전임교원_확보율_학과별.xlsx'
        → '대학_6-나-(1)_전임교원_확보율_학과별'
    """
    stem = Path(filename).stem
    stem = YEAR_IN_FILENAME.sub("", stem)   # 앞 연도 제거
    stem = re.sub(r"^[_\s]+", "", stem)     # 앞 밑줄/공백 제거
    return stem


def pub_year_from_filename(filename: str) -> int | None:
    """파일명 앞 연도를 공시연도로 반환.
    예) '2025년__대학_4-사_중도탈락...' → 2025
    """
    m = re.match(r"^(\d{4})년", Path(filename).name)
    return int(m.group(1)) if m else None


# ─────────────────────────────────────────────────────
# 필드 매핑 관리
# ─────────────────────────────────────────────────────

def parse_학교_field(df: "pd.DataFrame") -> "pd.DataFrame":
    """
    '학교' 컬럼을 파싱하여 '대학명', '본분교', '캠퍼스' 컬럼을 추가한다.

    학교 필드 규칙:
      {대학명} _제N캠퍼스  →  본분교=본교, 캠퍼스=제N캠퍼스
      {대학명} _분교       →  본분교=분교, 캠퍼스=
      {대학명}             →  본분교=본교, 캠퍼스=
    구분자: ' _' (공백 + 언더스코어)
    """
    if '학교' not in df.columns:
        return df

    def _parse(val: str):
        if not isinstance(val, str):
            return val, '본교', ''
        parts = val.split(' _', 1)
        대학명 = parts[0].strip()
        if len(parts) == 1:
            return 대학명, '본교', ''
        suffix = parts[1].strip()
        if suffix == '분교':
            return 대학명, '분교', ''
        return 대학명, '본교', suffix   # 제N캠퍼스 등

    parsed = df['학교'].apply(_parse)
    idx = df.columns.get_loc('학교') + 1   # '학교' 컬럼 바로 뒤에 삽입
    df.insert(idx,     '대학명', [r[0] for r in parsed])
    df.insert(idx + 1, '본분교', [r[1] for r in parsed])
    df.insert(idx + 2, '캠퍼스', [r[2] for r in parsed])
    return df


def load_mapping() -> dict:
    """field_mapping.json 로드.
    구조: { "__shared": {표준명: [alias...]}, "항목키": {표준명: [alias...]} }
    구버전 flat 구조(값이 list)는 자동으로 __shared로 마이그레이션.
    """
    if not MAPPING_FILE.exists():
        return {"__shared": {}}
    data = json.loads(MAPPING_FILE.read_text(encoding="utf-8"))
    # 구버전 감지: 최상위 값이 list → flat 구조
    if data and any(isinstance(v, list) for v in data.values()):
        data = {"__shared": data}
        MAPPING_FILE.write_text(
            json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return data


def save_mapping(mapping: dict):
    MAPPING_FILE.write_text(
        json.dumps(mapping, ensure_ascii=False, indent=2), encoding="utf-8")


def resolve_fields(raw_headers: list, mapping: dict, item_key: str) -> tuple:
    """
    raw_headers 각각을 표준 필드명으로 변환.
    - __shared 섹션 + item_key 섹션을 합산해 alias 역방향 맵 구성
    - 기존 매핑에 있으면 → 표준명으로 변환
    - 없으면 → (new_fields 목록에 추가, 원래 이름 그대로)
    Returns:
        resolved   : 표준 필드명 리스트
        new_fields : 매핑에 없는 새 필드명 리스트
    """
    shared   = mapping.get("__shared", {})
    item_sec = mapping.get(item_key, {})
    merged   = {**shared, **item_sec}   # item_key 섹션이 shared를 덮어씀

    alias_to_std = {}
    for std, aliases in merged.items():
        for alias in aliases:
            alias_to_std[alias] = std

    resolved   = []
    new_fields = []
    for h in raw_headers:
        if h in alias_to_std:
            resolved.append(alias_to_std[h])
        elif h in merged:               # 이미 표준명 자체
            resolved.append(h)
        else:
            resolved.append(h)
            new_fields.append(h)

    return resolved, new_fields


# ─────────────────────────────────────────────────────
# JSON 누적
# ─────────────────────────────────────────────────────

def accumulate_json(item_key: str, df: pd.DataFrame):
    """항목별 JSON 파일에 df 데이터를 연도 기준으로 누적."""
    JSON_DIR.mkdir(parents=True, exist_ok=True)
    json_path = JSON_DIR / f"{item_key}.json"

    # 기존 데이터 로드
    existing = []
    if json_path.exists():
        existing = json.loads(json_path.read_text(encoding="utf-8"))

    # 이번 파일의 공시연도 추출 → 같은 공시연도 기존 데이터 제거 (덮어쓰기)
    if '공시연도' in df.columns:
        new_pub_years = df['공시연도'].astype(str).unique().tolist()
        existing = [row for row in existing
                    if str(row.get("공시연도", row.get("기준연도", ""))) not in new_pub_years]
    else:
        # 구버전 호환: 공시연도 없는 df는 기준연도 기반 제거
        year_col  = df.columns[0]
        new_years = df[year_col].astype(str).unique().tolist()
        existing  = [row for row in existing if str(row.get("기준연도", "")) not in new_years]

    # 새 데이터 append
    new_records = df.where(pd.notnull(df), None).to_dict(orient="records")
    combined    = existing + new_records

    # 연도 제한 없이 전체 누적 보관
    # (표시 범위는 분석 페이지에서 제어)

    json_path.write_text(
        json.dumps(combined, ensure_ascii=False, indent=2), encoding="utf-8")

    return len(new_records), len(combined)


def validate_df(df: pd.DataFrame) -> list[str]:
    """정제 후 df의 데이터 이상 여부를 검사하여 경고 메시지 목록을 반환."""
    warnings: list[str] = []

    if df.empty:
        warnings.append("데이터 행이 없습니다.")
        return warnings

    # 기준연도 검사
    year_col = df.columns[0]
    bad_years = df[~df[year_col].astype(str).str.match(r"^(19|20)\d{2}$")][year_col]
    if not bad_years.empty:
        unique_bad = bad_years.dropna().unique()[:5]
        warnings.append(f"기준연도 이상값 {len(bad_years)}개: {list(unique_bad)}")

    # 공시연도 존재 여부 검사
    if '공시연도' not in df.columns:
        warnings.append("공시연도 컬럼이 없습니다 (파일명에서 연도 추출 실패)")

    # 대학명 / 학교 공백 검사
    for col in ("대학명", "학교"):
        if col in df.columns:
            null_cnt = df[col].isna().sum() + (df[col].astype(str).str.strip() == "").sum()
            if null_cnt:
                warnings.append(f"'{col}' 빈 값 {null_cnt}개")

    # 전체 행 대비 NaN 비율이 50% 초과인 컬럼
    threshold = len(df) * 0.5
    numeric_cols = [c for c in df.columns if c not in ("기준연도", "학교", "대학명", "본분교", "캠퍼스", "학과명")]
    for col in numeric_cols:
        if df[col].isna().sum() > threshold:
            warnings.append(f"'{col}' NaN {df[col].isna().sum()}/{len(df)}행 (50% 초과)")

    return warnings


def accumulate_csv(item_key: str, df: pd.DataFrame, output_dir: str) -> tuple[int, int]:
    """항목별 CSV 파일에 df 데이터를 연도 기준으로 누적 저장."""
    os.makedirs(output_dir, exist_ok=True)
    csv_path = os.path.join(output_dir, f"{item_key}.csv")

    year_col = df.columns[0]  # 첫 번째 컬럼이 기준연도
    new_years = df[year_col].astype(str).unique().tolist()

    # 기존 CSV 로드 후 같은 연도 제거 (덮어쓰기)
    existing_df = pd.DataFrame()
    if os.path.exists(csv_path):
        try:
            existing_df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str)
            existing_df = existing_df[~existing_df[year_col].astype(str).isin(new_years)]
        except Exception:
            existing_df = pd.DataFrame()

    combined = pd.concat([existing_df, df], ignore_index=True)
    combined.to_csv(csv_path, index=False, encoding="utf-8-sig")

    return len(df), len(combined)


# ─────────────────────────────────────────────────────
# 필드 매핑 확인 팝업
# ─────────────────────────────────────────────────────

class FieldMappingDialog(tk.Toplevel):
    """
    새 필드 또는 이름이 바뀐 필드를 사람이 확인하는 팝업.
    각 필드에 대해:
      - 기존 표준 필드에 매핑  (드롭다운)
      - 새 필드로 추가
      - 이 파일에서만 무시
    """
    def __init__(self, parent, new_fields: list, existing_std_fields: list, filename: str,
                 all_raw_headers: list = None):
        super().__init__(parent)
        self.title("필드 확인 필요")
        self.geometry("700x500")
        self.resizable(True, True)
        self.grab_set()   # 모달

        self.result   = None   # "ok" or "cancel"
        self.decisions = {}    # field → ("map", std_name) | ("add",) | ("ignore",)

        # 매핑 후보: 기존 표준 필드 + 이 파일의 이미 인식된 컬럼명 (중복 제거, 정렬)
        extra = [h for h in (all_raw_headers or []) if h not in new_fields and h not in existing_std_fields]
        combined_std = sorted(set(existing_std_fields + extra))

        self._build(new_fields, combined_std, filename)

    def _build(self, new_fields, existing_std_fields, filename):
        # 안내
        ttk.Label(self, text=f"파일: {filename}", font=("", 9, "bold")).pack(
            anchor="w", padx=16, pady=(12, 2))
        ttk.Label(self,
            text="기존 매핑에 없는 필드가 발견됐습니다.\n"
                 "각 필드를 기존 표준 필드에 매핑하거나, 새 필드로 추가하거나, 무시할 수 있습니다.",
            foreground="#555").pack(anchor="w", padx=16, pady=(0, 8))

        ttk.Separator(self).pack(fill="x", padx=16)

        # 스크롤 영역
        canvas = tk.Canvas(self, borderwidth=0, background="#fafafa")
        vscroll = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vscroll.set)
        vscroll.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True, padx=(16, 0), pady=8)

        frame = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=frame, anchor="nw")
        frame.bind("<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")))

        # 헤더
        for col, (text, w) in enumerate([("새 필드명", 280), ("처리 방식", 120), ("매핑 대상 (기존 필드)", 220)]):
            ttk.Label(frame, text=text, font=("", 9, "bold"), width=w//8).grid(
                row=0, column=col, padx=6, pady=4, sticky="w")

        self._rows = []
        options_map    = ["기존 필드에 매핑"]
        options_action = ["기존 필드에 매핑", "새 필드로 추가", "이 파일만 무시"]

        for i, field in enumerate(new_fields, start=1):
            # 필드명
            ttk.Label(frame, text=field, wraplength=270).grid(
                row=i, column=0, padx=6, pady=4, sticky="w")

            # 처리 방식 드롭다운
            var_action = tk.StringVar(value="새 필드로 추가")
            cb_action  = ttk.Combobox(frame, textvariable=var_action,
                                       values=options_action, state="readonly", width=16)
            cb_action.grid(row=i, column=1, padx=6, pady=4)

            # 매핑 대상 드롭다운 (자유 입력 허용: 목록이 비어도 직접 타이핑 가능)
            var_target = tk.StringVar(value="")
            cb_target  = ttk.Combobox(frame, textvariable=var_target,
                                       values=existing_std_fields, state="disabled", width=26)
            cb_target.grid(row=i, column=2, padx=6, pady=4)

            # 처리 방식 변경 시 매핑 대상 활성/비활성
            def on_action_change(event, ct=cb_target, va=var_action):
                ct.configure(state="normal" if va.get() == "기존 필드에 매핑" else "disabled")
                if va.get() != "기존 필드에 매핑":
                    ct.set("")

            cb_action.bind("<<ComboboxSelected>>", on_action_change)
            cb_target.configure(state="disabled")

            self._rows.append((field, var_action, var_target))

        # 버튼
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", padx=16, pady=12)
        ttk.Button(btn_frame, text="확인 → 계속 처리", command=self._ok).pack(side="left")
        ttk.Button(btn_frame, text="취소 (이 파일 건너뜀)", command=self._cancel).pack(
            side="left", padx=8)

    def _ok(self):
        decisions = {}
        for field, var_action, var_target in self._rows:
            action = var_action.get()
            if action == "기존 필드에 매핑":
                target = var_target.get().strip()
                if not target:
                    messagebox.showwarning("입력 필요",
                        f"'{field}' 필드의 매핑 대상을 선택하거나\n처리 방식을 변경하세요.")
                    return
                decisions[field] = ("map", target)
            elif action == "새 필드로 추가":
                decisions[field] = ("add",)
            else:
                decisions[field] = ("ignore",)

        self.decisions = decisions
        self.result    = "ok"
        self.destroy()

    def _cancel(self):
        self.result = "cancel"
        self.destroy()


# ─────────────────────────────────────────────────────
# 메인 GUI
# ─────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("대학알리미 공시 파일 정제")
        self.geometry("820x680")
        self.resizable(True, True)
        self.configure(bg="#f5f5f5")
        self._build_ui()

    def _build_ui(self):
        pad = {"padx": 16, "pady": 6}

        # ── 입력 폴더 ──
        fi = ttk.LabelFrame(self, text="입력 폴더 (xlsx 파일 위치)")
        fi.pack(fill="x", **pad)
        self.var_input = tk.StringVar(value=str(Path.home() / "Downloads"))
        ttk.Entry(fi, textvariable=self.var_input).pack(
            side="left", padx=8, pady=6, fill="x", expand=True)
        ttk.Button(fi, text="폴더 선택", command=self._pick_input).pack(
            side="left", padx=(0, 8), pady=6)

        # ── 출력 폴더 ──
        fo = ttk.LabelFrame(self, text="출력 폴더 (CSV / Excel 저장 위치)")
        fo.pack(fill="x", **pad)
        self.var_output = tk.StringVar(
            value=str(Path.home() / "Downloads" / "대학알리미_정제"))
        ttk.Entry(fo, textvariable=self.var_output).pack(
            side="left", padx=8, pady=6, fill="x", expand=True)
        ttk.Button(fo, text="폴더 선택", command=self._pick_output).pack(
            side="left", padx=(0, 8), pady=6)

        # ── 처리 옵션 ──
        fopt = ttk.LabelFrame(self, text="처리 옵션")
        fopt.pack(fill="x", **pad)

        self.var_csv      = tk.BooleanVar(value=True)
        self.var_xlsx_out = tk.BooleanVar(value=False)
        self.var_json_acc = tk.BooleanVar(value=True)

        ttk.Checkbutton(fopt, text="CSV 저장  (검증용)",
                        variable=self.var_csv).pack(side="left", padx=16, pady=6)
        ttk.Checkbutton(fopt, text="Excel 저장  (검증용)",
                        variable=self.var_xlsx_out).pack(side="left", padx=8, pady=6)
        ttk.Separator(fopt, orient="vertical").pack(side="left", fill="y", padx=8, pady=4)
        ttk.Checkbutton(fopt, text="항목별 JSON 누적  (분석툴용)",
                        variable=self.var_json_acc).pack(side="left", padx=8, pady=6)

        # JSON 저장 위치 표시
        ttk.Label(fopt, text=f"→ {JSON_DIR}", foreground="#777",
                  font=("", 8)).pack(side="left", pady=6)

        # ── 실행 버튼 ──
        fb = tk.Frame(self, bg="#f5f5f5")
        fb.pack(fill="x", padx=16, pady=4)
        self.btn_run = ttk.Button(fb, text="▶  정제 시작",
                                   command=self._run, width=16)
        self.btn_run.pack(side="left")
        ttk.Button(fb, text="📂  출력 폴더 열기",
                   command=self._open_output).pack(side="left", padx=8)
        ttk.Button(fb, text="📂  JSON 폴더 열기",
                   command=self._open_json_dir).pack(side="left")
        self.lbl_status = ttk.Label(fb, text="", foreground="#555")
        self.lbl_status.pack(side="left", padx=12)

        # ── 진행바 ──
        self.progress = ttk.Progressbar(self, mode="determinate")
        self.progress.pack(fill="x", padx=16, pady=(2, 4))

        # ── 로그 ──
        fl = ttk.LabelFrame(self, text="처리 로그")
        fl.pack(fill="both", expand=True, **pad)
        self.log = tk.Text(fl, wrap="word", state="disabled",
                            bg="white", fg="#222", font=("Courier", 10))
        sc = ttk.Scrollbar(fl, command=self.log.yview)
        self.log.configure(yscrollcommand=sc.set)
        sc.pack(side="right", fill="y")
        self.log.pack(fill="both", expand=True, padx=4, pady=4)
        self.log.tag_config("ok",    foreground="#1a7a2e")
        self.log.tag_config("warn",  foreground="#d35400")
        self.log.tag_config("error", foreground="#c0392b")
        self.log.tag_config("skip",  foreground="#888888")
        self.log.tag_config("info",  foreground="#1a5276")
        self.log.tag_config("bold",  font=("Courier", 10, "bold"))

    # ── 폴더 선택 ──
    def _pick_input(self):
        d = filedialog.askdirectory(initialdir=self.var_input.get())
        if d:
            self.var_input.set(d)
            self.var_output.set(str(Path(d) / "대학알리미_정제"))

    def _pick_output(self):
        d = filedialog.askdirectory(initialdir=self.var_output.get())
        if d:
            self.var_output.set(d)

    def _open_dir(self, path):
        import subprocess, sys
        if not os.path.exists(path):
            messagebox.showinfo("알림", f"폴더가 없습니다:\n{path}")
            return
        if sys.platform == "win32":
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.run(["open", path])
        else:
            subprocess.run(["xdg-open", path])

    def _open_output(self):
        self._open_dir(self.var_output.get())

    def _open_json_dir(self):
        self._open_dir(str(JSON_DIR))

    # ── 로그 출력 ──
    def _log(self, msg, tag=""):
        self.log.configure(state="normal")
        self.log.insert("end", msg + "\n", tag)
        self.log.see("end")
        self.log.configure(state="disabled")

    # ── 실행 ──
    def _run(self):
        input_dir  = self.var_input.get()
        output_dir = self.var_output.get()

        xlsx_files = sorted(Path(input_dir).glob("*.xlsx"))
        if not xlsx_files:
            messagebox.showwarning("파일 없음",
                f"{input_dir}\n폴더에 xlsx 파일이 없습니다.")
            return

        if not self.var_csv.get() and not self.var_xlsx_out.get() \
                and not self.var_json_acc.get():
            messagebox.showwarning("옵션 미선택", "처리 옵션을 하나 이상 선택하세요.")
            return

        self.btn_run.configure(state="disabled")
        self.log.configure(state="normal")
        self.log.delete("1.0", "end")
        self.log.configure(state="disabled")

        threading.Thread(
            target=self._process,
            args=(xlsx_files, output_dir),
            daemon=True
        ).start()

    def _process(self, files, output_dir):
        total   = len(files)
        mapping = load_mapping()

        self.progress["maximum"] = total
        self.progress["value"]   = 0
        self._log(f"총 {total}개 파일 처리 시작", "info")
        self._log("─" * 64)

        ok = skip = error = 0

        for i, filepath in enumerate(files):
            fname = filepath.name
            self._log(f"[{i+1}/{total}] {fname}", "bold")

            try:
                # 1. 정제
                df, sheet_name, header_rows, data_start = extract_df(str(filepath))
                raw_headers = list(df.columns)

                # 1-1. 항목 키 먼저 확정 (resolve_fields에 전달 필요)
                item_key_early = item_key_from_filename(fname)

                # 2. 필드 매핑 적용
                resolved, new_fields = resolve_fields(raw_headers, mapping, item_key_early)

                # 3. 새 필드 있으면 메인 스레드에서 팝업
                if new_fields:
                    self._log(f"  ⚠️  새 필드 {len(new_fields)}개 발견: "
                              f"{new_fields[:3]}{'...' if len(new_fields)>3 else ''}", "warn")

                    # 팝업 후보: shared + 이 항목 섹션의 표준 필드명
                    shared_keys   = list(mapping.get("__shared", {}).keys())
                    item_keys_now = list(mapping.get(item_key_early, {}).keys())
                    existing_std  = sorted(set(shared_keys + item_keys_now))
                    decisions    = self._ask_mapping(new_fields, existing_std, fname,
                                                     all_raw_headers=raw_headers)

                    if decisions is None:
                        self._log("  ⏭  사용자가 건너뜀", "skip")
                        skip += 1
                        self.progress["value"] = i + 1
                        continue

                    # 결정 반영 — 새 필드는 항목 섹션에 저장
                    item_sec   = mapping.setdefault(item_key_early, {})
                    shared_sec = mapping.setdefault("__shared", {})

                    for field, decision in decisions.items():
                        if decision[0] == "map":
                            std = decision[1]
                            # std가 속한 섹션에 alias 추가
                            if std in shared_sec:
                                target = shared_sec
                            else:
                                target = item_sec
                                item_sec.setdefault(std, [])
                            if field not in target[std]:
                                target[std].append(field)
                            resolved = [std if r == field else r for r in resolved]
                            self._log(f"    매핑: '{field}' → '{std}'", "info")

                        elif decision[0] == "add":
                            # 이 항목 섹션에 새 표준 필드 등록
                            if field not in item_sec and field not in shared_sec:
                                item_sec[field] = []
                            self._log(f"    새 필드 추가: '{field}'", "info")

                        else:  # ignore
                            idx_to_drop = [j for j, h in enumerate(raw_headers) if h == field]
                            df.drop(df.columns[idx_to_drop], axis=1, inplace=True,
                                    errors="ignore")
                            resolved = [r for r, h in zip(resolved, raw_headers)
                                        if h != field]
                            self._log(f"    무시: '{field}'", "skip")

                    save_mapping(mapping)

                # 4. 컬럼명 교체
                df = df.iloc[:, :len(resolved)]
                df.columns = resolved

                # 4-1. '학교' 필드 파싱 → 대학명 / 본분교 / 캠퍼스 추가
                df = parse_학교_field(df)

                # 4-1-1. 반기 표기 정규화 (재학생 충원율 등 상/하반기 항목)
                before = len(df)
                df = normalize_half_year(df)
                dropped = before - len(df)
                if dropped:
                    self._log(f"  ℹ️  하반기 행 {dropped}개 제거 (상반기만 보관)", "info")

                # 4-2. 공시연도 삽입 (파일명 앞 연도)
                pub_year = pub_year_from_filename(fname)
                if pub_year is not None:
                    df.insert(1, '공시연도', pub_year)
                else:
                    self._log("  ⚠️  파일명에서 공시연도를 추출할 수 없습니다.", "warn")

                item_key = item_key_early   # 앞서 계산한 값 재사용

                # 4-2. 데이터 검증
                for warn_msg in validate_df(df):
                    self._log(f"  ⚠️  검증 경고: {warn_msg}", "warn")

                # 5. CSV 저장 (연도별 누적)
                if self.var_csv.get():
                    csv_new, csv_total = accumulate_csv(item_key, df, output_dir)

                # 6. Excel 저장
                if self.var_xlsx_out.get():
                    os.makedirs(output_dir, exist_ok=True)
                    xl_path = os.path.join(output_dir, f"{filepath.stem}_정제.xlsx")
                    df.to_excel(xl_path, index=False)

                # 7. JSON 누적
                csv_msg = ""
                if self.var_csv.get():
                    csv_msg = f" | CSV +{csv_new}행 (누적 {csv_total}행)"

                json_msg = ""
                if self.var_json_acc.get():
                    new_cnt, total_cnt = accumulate_json(item_key, df)
                    json_msg = f" | JSON +{new_cnt}행 (누적 {total_cnt}행)"

                saved_fmts = []
                if self.var_csv.get():      saved_fmts.append("CSV누적")
                if self.var_xlsx_out.get(): saved_fmts.append("Excel")
                if self.var_json_acc.get(): saved_fmts.append("JSON누적")

                self._log(
                    f"  ✅ {len(df)}행 × {len(df.columns)}열"
                    f" | 시트: {sheet_name}"
                    f" | 헤더: {header_rows}행"
                    f"{csv_msg}{json_msg}", "ok")
                ok += 1

            except Exception as e:
                self._log(f"  ❌ 오류: {e}", "error")
                error += 1

            self.progress["value"] = i + 1
            self.lbl_status.configure(
                text=f"{i+1} / {total}  (✅{ok}  ❌{error}  ⏭{skip})")

        self._log("─" * 64)
        self._log(f"완료: {ok}개 성공  /  {error}개 오류  /  {skip}개 스킵", "bold")
        self.btn_run.configure(state="normal")

    def _ask_mapping(self, new_fields, existing_std, filename, all_raw_headers=None) -> dict | None:
        """메인 스레드에서 필드 매핑 팝업 실행 후 결과 반환"""
        result_holder = [None]

        def show():
            dlg = FieldMappingDialog(self, new_fields, existing_std, filename,
                                     all_raw_headers=all_raw_headers)
            self.wait_window(dlg)
            result_holder[0] = (dlg.result, dlg.decisions)

        self.after(0, show)

        # 팝업이 닫힐 때까지 대기 (스레드에서 폴링)
        import time
        while result_holder[0] is None:
            time.sleep(0.05)

        result, decisions = result_holder[0]
        if result == "cancel":
            return None
        return decisions


if __name__ == "__main__":
    app = App()
    app.mainloop()
